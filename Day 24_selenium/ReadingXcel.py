import openpyxl

#File --> Workbook--> Sheets-->Rows--> Cells

file = "C:\selenium\Data.xlsx"
workbook = openpyxl.load_workbook(file)
sheet= workbook["Sheet1"]

rows = sheet.max_row    # count no. of rows in excel sheet
cols= sheet.max_column   # count no. of colums in excel sheet

#print(rows,cols)

#reading all the columns and rows from xcel using for loop
for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end= '            ')
    print()

