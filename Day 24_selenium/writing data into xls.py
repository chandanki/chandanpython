import openpyxl
#File --> Workbook--> Sheets-->Rows--> Cells

#(for same data)
# file = "C:\selenium\Data1.xlsx"
# workbook = openpyxl.load_workbook(file)
# sheet= workbook["Sheet1"] # or sheet = workbook.active
#
# for r in range(1,5):
#     for c in range(1,4):
#         sheet.cell(r,c).value="chandan"
#
# workbook.save(file)

#(for different and multiple data)
file = "C:\selenium\Data1.xlsx"
workbook = openpyxl.load_workbook(file)
sheet= workbook.active  # or sheet = workbook.active

sheet.cell(1,1).value="ID"
sheet.cell(1,2).value="Name"
sheet.cell(1,3).value="sal"

sheet.cell(2,1).value= 1
sheet.cell(2,2).value="Chandan"
sheet.cell(2,3).value=5000

sheet.cell(3,1).value=2
sheet.cell(3,2).value="john"
sheet.cell(3,3).value=6000

sheet.cell(4,1).value=3
sheet.cell(4,2).value="champ"
sheet.cell(4,3).value=8000

workbook.save(file)

