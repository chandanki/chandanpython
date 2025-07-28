import openpyxl

file = open("file path","w")
file.write("this is my first line\n")
file.close()

# read file
file = open("file path","r")
print(file.read())
file.close()

# Read and write in selenium
file = "file path"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active #workbook[sheet 1]
for r in range(1,5):
    for c in range(1,4):
        sheet.cell(r,c).value = "Text"
workbook.save(file)

# Read
file = "file Path"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active
for r in range (1,5):
    for c in range (1,6):
        print(sheet.cell(r,c).value)
print()
